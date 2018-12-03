# -*- coding: utf-8 -*-
"""
Created on Tue Oct  2 14:22:30 2018

D:\Tools\Python_tools\Python_examples\Python_examples\py_3doeAPI_json_ex1.py

作为一个3doe公司efoot的API演示程序

请联系我们提供url的服务器名字 


@author: Wisdom Zhang
"""

import json, requests, sys
import re

import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


url='http://localhost/pm8_efoot_v3/getfootphone3/pm8_efoot_v3/getfootphone3.php?token='
token='0e90c5138a24797d6b57914cd2687ac1'
ph='0450530722'

# Download the JSON data from OpenWeatherMap.org's API.
url =url + (token)+'&ph='+ph
response = requests.get(url)
response.raise_for_status()
# TODO: Load JSON data into a Python variable.

#eFootAPIData = unicode( response.text, errors='ignore') 
# Load JSON data into a Python variable.
eFootData = json.loads(response.text[3:])

# 2nd method get the data
eFootAPIData=re.sub('\'','\"',response.text)

result=json.loads(eFootAPIData[3:])

for efoot in eFootData['data']['customerfoot']:
    print(efoot)
    for footpara,footvalue in efoot.items():
        print(footpara,':',footvalue)
        
efoot_R=eFootData['data']['customerfoot'][1]
efoot_L=eFootData['data']['customerfoot'][0]
# open the excel


wb2 = load_workbook('foot_report.xlsx')
print (wb2.get_sheet_names())

for sheet in wb2:
    print(sheet.title)


for sheet in wb2:
    print(sheet.title)

# the foot report sheet
ws3 = wb2["Sheet1"]

ws3['A1']= datetime.datetime(2010,7,21)
ws3['A1'].number_format

ws3['A2']='API'
ws3['B1'] = 'this data from 3doe api '
wb2.guess_types = False
ws3['B1'].value
ws3['B1'].number_format


ws3['C4']=efoot_L['phone'] 
ws3['C3']=efoot_L['user_name']
ws3['F2']=efoot_L['ScanTime'][0:9]
ws3['I2']=efoot_L['ScanTime'][10:]
ws3['K2']=efoot_L['store_name']
# Save the file
wb2.save("foot_report_api.xlsx")
#wb2.close()


